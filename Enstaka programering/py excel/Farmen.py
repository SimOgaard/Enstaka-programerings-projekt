import openpyxl
from collections import Counter

wb = openpyxl.load_workbook('Farmen.xlsx')

Deltagar_sheet = wb.get_sheet_by_name('Deltagare')
Veckosammanfattning_sheet = wb.get_sheet_by_name('Veckosammanfattning')
Tittarsiffror_sheet = wb.get_sheet_by_name('Tittarsiffror')

deltagaren = [Deltagar_sheet.cell(row=i, column=1).value for i in range(1, Deltagar_sheet.max_row)]

deltagare = [{"Namn":Deltagar_sheet.cell(row=i,column=1).value,
        "År":Deltagar_sheet.cell(row=i,column=2).value,
        "Bostad":Deltagar_sheet.cell(row=i,column=3).value,
        "Syssels ättning":Deltagar_sheet.cell(row=i,column=4).value,
        "Resultat":Deltagar_sheet.cell(row=i,column=5).value}
for i in range(2, Deltagar_sheet.max_row)]

[print(deltagaren["Namn"], "är", deltagaren["År"], "år gammal och kommer från", deltagaren["Bostad"], "där hens syssel sättning är", deltagaren["Syssels ättning"]) for deltagaren in deltagare]

vecka = [Veckosammanfattning_sheet.cell(row=i, column=2).value for i in range (2, Veckosammanfattning_sheet.max_row)]

veckan = [{"Vecka":Veckosammanfattning_sheet.cell(row=i,column=1).value,
        "Storbonde":Veckosammanfattning_sheet.cell(row=i,column=2).value,
        "Förstekämpe":Veckosammanfattning_sheet.cell(row=i,column=3).value,
        "Andrekämpe":Veckosammanfattning_sheet.cell(row=i,column=4).value,
        "Tvekamp":Veckosammanfattning_sheet.cell(row=i,column=5).value,
        "Hemskickad":Veckosammanfattning_sheet.cell(row=i,column=6).value}
for i in range(2, Veckosammanfattning_sheet.max_row)]

[print("under vecka", vecka["Vecka"], "var", vecka["Storbonde"], "Storbonde med sinna två kämpar", vecka["Förstekämpe"], "och", vecka["Andrekämpe"], "de tre körde tvekampen", vecka["Tvekamp"], "där", vecka["Hemskickad"], "blev hemskickad. Denna vecka fick", vecka["Hemskickad"], "visningar") for vecka in veckan]

vecka2 = [Tittarsiffror_sheet.cell(row=i, column=1).value for i in range (1, Veckosammanfattning_sheet.max_row)]

veckan2 = [{"Vecka":Tittarsiffror_sheet.cell(row=i,column=1).value,
        "Avsnitt":Veckosammanfattning_sheet.cell(row=i,column=2).value,
        "Datum":Veckosammanfattning_sheet.cell(row=i,column=3).value,
        "Andrekämpe":Veckosammanfattning_sheet.cell(row=i,column=4).value,
        "Tvekamp":Veckosammanfattning_sheet.cell(row=i,column=5).value,
        "Hemskickad":Veckosammanfattning_sheet.cell(row=i,column=6).value}
for i in range(2, Veckosammanfattning_sheet.max_row)]


tävlademest = [Veckosammanfattning_sheet.cell(row=i, column=q).value for q in range(2,5) for  i in range (2, Veckosammanfattning_sheet.max_row)]
print(tävlademest)
räknad = Counter(tävlademest)
print(räknad)

popisgrenen = [Veckosammanfattning_sheet.cell(row=i, column=5).value for i in range (2, Veckosammanfattning_sheet.max_row)]
räknad = Counter(popisgrenen)
# mest=
# mest=mest[0:1]
# print(mest)
print(räknad)

hemskickad = [Veckosammanfattning_sheet.cell(row=i, column=6).value for i in range(2, Veckosammanfattning_sheet.max_row)]
räknad = Counter(hemskickad)
print(räknad)