import openpyxl
wb = openpyxl.load_workbook('mello.xlsx')
sheet = wb.active

print(sheet['A1'].value)

print(sheet.cell(row=1, column=1).value)

värden=[i for i in range(2,10)]
print(värden)

värden=["Värde "+str(i) for i in range(2,10) if i%2 is 0]
print(värden)

värden=["Värde "+str(i) for i in range(2,10) if i%2 is 0 if i is not 8]
print(värden)

artist=[]
for i in range(2,sheet.max_row):
    artist.append(sheet.cell(row=i, column=2).value)
print(artist)

artist=[sheet.cell(row=i, column=2).value for i in range(2,sheet.max_row)]
print(artist)

Nano = {'Namn': sheet.cell(row=2, column=2).value, 'Sång': sheet.cell(row=2, column=3).value, 'Poäng': sheet.cell(row=2, column=6).value, 'Röst': sheet.cell(row=2, column=5).value}
print(Nano)

artists=[{"Namn":sheet.cell(row=i,column=2).value,
        "Sång":sheet.cell(row=i,column=3).value,
        "Poäng":sheet.cell(row=i,column=6).value,
        "Röst":sheet.cell(row=i,column=5).value}
for i in range(2,sheet.max_row)]
print(artists)

vidare=[{"Namn":sheet.cell(row=i,column=2).value,
        "Sång":sheet.cell(row=i,column=3).value,
        "Poäng":sheet.cell(row=i,column=6).value,
        "Röst":sheet.cell(row=i,column=5).value}
for i in range(2,sheet.max_row)
    if not sheet.cell(row=i,column=8).value.startswith("Utsl")]
print(vidare)

[print(artist["Namn"]," är med i år!") for artist in artists]
[print(artist["Namn"], "var en av de",len(vidare)," som gick vidare med låten", artist["Sång"]) for artist in vidare]

vidare = sorted(vidare, key=lambda k: k['Poäng'])
vidare = sorted(vidare, key=lambda k: k['Poäng'], reverse=True)

vidare = vidare[0:10]

[print(artist["Namn"]," är med i år!") for artist in artists]
[print(artist["Namn"], "var en av de",len(vidare)," som gick vidare med låten", artist["Sång"]) for artist in vidare]